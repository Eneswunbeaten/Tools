import os
import asyncio
from openpyxl import load_workbook
from googletrans import Translator

async def translate_excel():
    print("Welcome to the Excel Translator!")
    print("Please enter the full path to the Excel file you want to translate (e.g., C:/Users/Name/Desktop/file.xlsx):")
    excel_path = input("Excel file path: ")
    print("Please enter the source language ISO code (e.g., sk for Slovak, en for English, tr for Turkish). If you are not sure, you can type 'auto' for automatic detection:")
    src_lang = input("Source language code: ")
    print("Please enter the target language ISO code (e.g., tr for Turkish, en for English, sk for Slovak):")
    dest_lang = input("Target language code: ")

    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        translator = Translator()
        print("Translation is starting. Please wait...")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if not cell.value or not isinstance(cell.value, str):
                        continue
                    if cell.value.strip().startswith("="):
                        continue
                    try:
                        translation = await translator.translate(cell.value, src=src_lang, dest=dest_lang)
                        print(f"{cell.coordinate}: {cell.value} -> {translation.text}")
                        cell.value = translation.text
                    except Exception as e:
                        print(f"Cell {cell.coordinate} could not be translated: {e}")

        directory, file_name = os.path.split(excel_path)
        new_file_name = f"translated-{file_name}"
        save_path = os.path.join(directory, new_file_name)
        wb.save(save_path)
        print(f"Translation completed! The translated Excel file has been saved as '{save_path}'.")
    else:
        print(f"Error: The file path '{excel_path}' does not exist or is not valid.")

if __name__ == "__main__":
    asyncio.run(translate_excel())
